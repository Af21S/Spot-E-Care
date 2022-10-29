##########################         $pot-E-Care      ############################
from tkinter import *
import openpyxl
from openpyxl.styles import Alignment, Font
label=0
user_details=openpyxl.load_workbook("User_details.xlsx")
billBook=openpyxl.load_workbook('BillBook.xlsx')
usersheet=user_details.active
userCart={}
item_price={'Zoho':1610,"Tally":7204,"Adobe Photoshop CC for team":7000,
                "Dr.Fone-Data Recovery":3000,"Vyapar Billing Software":724,
                "Microsoft Teams":1500,"Grammerly":1000,"Adobe Acrobat Pro DC for Teams":2000,
                "Microsoft Office 365":2000,"Wise Data Recovery":6000,"CCleaner":1600,
                "Krisp":7680,"F.lux":800,"Revo Uninstaller Pro":3120,"iA Writer":2320,
                "Breevy":2720,"Fences":880,"Microsoft Office":3000,"Quick Heal":824,"Adobe Lightroom CC":6500,
                "Stellar Data Recovery":5000}
global bs
bs=[]
bs.append(billBook.sheetnames[0])
bs[0]=billBook.active
global username
username=''
global totalPrice



def first() :
    print("----------------  Welcome to $pot-E-Care !! -----------------")
    print("Enter \n1.Start Shopping\n2. User Credentials\n3.Exit")
    choice=int(input())
    if choice==1 :
        Login()
    elif choice==2 :
        till=False
        while till!=True :
            list1=checkPerson()
            flag=list1[0]
            person=list1[1]
            if flag==1 :
                print(f"USERNAME:      {usersheet['a'+str(person)].value}")
                print(f"PHONE NUMBER:  {usersheet['c'+str(person)].value}")
                print(f"EMAIL-ID:      {usersheet['d'+str(person)].value}")
                print(f"GENDER:        {usersheet['e'+str(person)].value}")
                print(f"Age:           {usersheet['f'+str(person)].value}")
                till=True
            else :
                print("Invalid Username orPassword.")
                print("Enter\n1. Exit\n2. Try Again")
                ch=int(input())
                if ch==1 :
                    end()
    else :
        end()


def checkPerson() :
    x=len(usersheet['a'])
    flag=0
    row=0
    global username
    username=input("Enter Username: ")
    password=input("Enter Password: ")
    for person in range(2,x+1) :
        if username==usersheet['a'+str(person)].value :
            if password==usersheet['b'+str(person)].value :
                flag=1
                row=person
    return [flag,row]


def Login() :
    list1=checkPerson()
    flag=list1[0]
    if flag==1 :
        print("Login Successfull")
        cart()
    else :
        print("꒰⍨꒱")
        user()
        print("1 Login \n2. Sign UP \n3Exit")
        choice=int(input())
        if choice==1:
            Login()
        elif choice==2:
            signUp()
        else :
            end()


def signUp() :
    x=len(usersheet['a'])
    global username
    username=input('Enter Username: ')
    usersheet['a'+str(x+1)]=username
    password=input('Enter Password: ')
    usersheet['b'+str(x+1)]=password
    phone=int(input('Enter Phone number: '))
    usersheet['c'+str(x+1)]=phone
    email_id=input('Enter Email id: ')
    usersheet['d'+str(x+1)]=email_id
    gender=input('Enter Gender: ')
    usersheet['e'+str(x+1)]=gender
    age=input('Enter Age: ')
    usersheet['f'+str(x+1)]=age
    user_details.save("User_details.xlsx")
    print("Enter \n1. Login\n2.Exit")
    choice=int(input())
    if choice==1:
        Login()
    else :
        end()

        
def user():
    root = Tk()
    myLabel=Label(root,text="Namaste User!☻")
    myLabel.pack()
    myLabel=Label(root,text="Sorry, we couldn't find you")
    myLabel.pack()

    myLabel=Label(root,text="Please Try again")
    myLabel.pack()
    root.geometry("300x300")
    root.attributes("-topmost", True)
    def Close(): 
       root.destroy() 
    # Button for closing 

    try_button = Button(root, text="Try Again", command=Close) 
    try_button.pack(pady=20) 
    root.mainloop()



def cart() :
    till=False
    while till!=True :
        print("Enter\n1. Add Package(s)\n2. View Cart")
        choice=int(input())
        if choice==1 :
            itemsDetails()
            pack=int(input("Enter the Package code: "))
            while till!=True :
                duration=int(input("Enter validity(max 3years): "))
                if(duration>3):
                    print("Validity is maximum for 3 years")
                else :
                    till=True
            till=False
            userCart.update({list(item_price.keys())[pack-1]:[list(item_price.values())[pack-1],duration,list(item_price.values())[pack-1]*duration]}) 
        else :
            print(f"You have {len(userCart)} packages in cart.")
            print("CODE\tPACKAGE\t\t\t\tPRICE\tDURATION")
            for elem in range(len(userCart)) :
                len_packageName=len(list(item_price.keys())[elem])
                if len_packageName<=6 :
                    print(f" {elem+1} \t{list(userCart.keys())[elem]}\t\t\t\t{list(userCart.values())[elem][0]}\t{list(userCart.values())[elem][1]}")
                elif len_packageName<=15 :
                    print(f" {elem+1} \t{list(userCart.keys())[elem]}\t\t\t{list(userCart.values())[elem][0]}\t{list(userCart.values())[elem][1]}")
                elif len_packageName<=25 :
                    print(f" {elem+1} \t{list(userCart.keys())[elem]}\t\t{list(userCart.values())[elem][0]}\t{list(userCart.values())[elem][1]}")
                else :
                    print(f" {elem+1} \t{list(userCart.keys())[elem]}\t{list(userCart.values())[elem][0]}\t{list(userCart.values())[elem][1]}")
            flag=0
            print("Enter\n1. Bill\n2. Discard Package\n3. Continue")
            ch=int(input())
            if ch==1 :
                flag=1
                till=True
            elif ch==2 :
                delete()
    if flag==1 :
        bill()


def itemsDetails() :
    print("Code\tPACKAGES\t\t\t\tPRICE(per year)\n")
    for package in range(len(item_price)) :
        len_packageName=len(list(item_price.keys())[package])
        if len_packageName<=6 :
            print(f"{package+1}\t{list(item_price.keys())[package]}\t\t\t\t\tRs.{list(item_price.values())[package]}")
        elif len_packageName<=15 :
            print(f"{package+1}\t{list(item_price.keys())[package]}\t\t\t\tRs.{list(item_price.values())[package]}")
        elif len_packageName<=25 :
            print(f"{package+1}\t{list(item_price.keys())[package]}\t\t\tRs.{list(item_price.values())[package]}")
        else :
            print(f"{package+1}\t{list(item_price.keys())[package]}\t\tRs.{list(item_price.values())[package]}")


def delete():
    code=int(input("Enter the Package Code to be discarded: "))
    trash=userCart.pop(list(userCart.keys())[code-1])
    print("Dicarded successfully.")


def bill() :
    global totalPrice
    sheets=len(billBook.sheetnames)
    if sheets==1 :
        if bs[0].title!="BILL1" :#not filled
            bs[0].title='BILL1'
            designBill(0)
        else :#filled
            bs.append(billBook.create_sheet("BILL2"))
            designBill(sheets)
    else :
        bs.append(billBook.create_sheet("BILL"+str(sheets+1)))
        designBill(sheets)
    print(f'YOUR BILL: RS.{totalPrice}')
    end()


def designBill(sh_num):
    global totalPrice
    totalPrice=0
    page=bs[sh_num-len(billBook.sheetnames)]
    page.merge_cells('a1:e2')
    page['a1'].font=Font(size=16,bold=True,underline='double')
    page['a1'].alignment=Alignment(horizontal='center', vertical='center')
    page['a1']="BILL NO. "+str(sh_num+1)
    page['a3'].font=Font(bold=True,underline='single')
    page['a3'].alignment=Alignment(horizontal='center', vertical='center')
    page['a3']="NAME: "
    page['b3']=username
    page['c3'].font=Font(bold=True,underline='single')
    page['c3'].alignment=Alignment(horizontal='center', vertical='center')
    page['c3']='EMAIL-ID:'
    x=len(usersheet['a'])
    for person in range(2,x+1) :
        if username==usersheet['a'+str(person)].value :
            page['d3']=usersheet['d'+str(person)].value
    for row in 'abcde' :
        page[str(row)+'4'].font=Font(bold=True,underline='single')
        page[str(row)+'4'].alignment=Alignment(horizontal='center', vertical='center')
    page['a4']="SL. NO."
    page['b4']="PACKAGE"
    page.column_dimensions['b'].width=40
    page['c4']="DURATION(in yrs)"
    page.column_dimensions['c'].width=20
    page['d4']="PACKAGE PRICE(Rs. per year)"
    page.column_dimensions['d'].width=30
    
    page['e4']="PACKAGE COST(Rs.)"
    page.column_dimensions['e'].width=20
    for row in range(len(userCart)) :
        page['a'+str(row+5)]=row+1
        page['b'+str(row+5)]=list(userCart.keys())[row]
        page['c'+str(row+5)]=list(userCart.values())[row][1]
        page['D'+str(row+5)]=list(userCart.values())[row][0]
        page['E'+str(row+5)]=list(userCart.values())[row][2]
        totalPrice+=list(userCart.values())[row][2]
    page['d'+str(len(userCart)+5)].font=Font(bold=True,underline='single')
    page['d'+str(len(userCart)+5)].alignment=Alignment(horizontal='center', vertical='center')
    page['d'+str(len(userCart)+5)]='TOTAL PRICE(Rs.): '
    page['e'+str(len(userCart)+5)]=totalPrice
    billBook.save('BillBook.xlsx')


def end() :
    print("------------------ Thank you ------------------------")
    print("---------------- Shop Again !! ----------------------")
    exit


first()






























